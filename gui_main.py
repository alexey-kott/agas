# coding=utf-8
import re
import sys
from pathlib import Path
from typing import List, Tuple

from tkinter import Tk, Label, Entry, Button, mainloop, filedialog, END
from bs4 import BeautifulSoup
from bs4.element import Tag
from openpyxl import load_workbook


def select_report(title: str,
                  element: Entry) -> None:
    """Report -- it's an HTML-file"""
    filetypes = [('HTML files', '*.html'),
                 ('All files', '*.*')]
    file_names = filedialog.askopenfilenames(title=title, filetypes=filetypes)

    element.delete(0, END)
    element.insert(0, ';'.join(file_names))


def select_template(title: str,
                    element: Entry) -> None:
    """Template -- it's Excel-file"""
    filetypes = [('Excel files', '*.xlsx'),
                 ('All files', '*.*')]
    file_names = filedialog.askopenfilename(title=title, filetypes=filetypes)
    element.delete(0, END)
    element.insert(0, file_names)


def is_interesting(tr: Tag, azs_prefixes: List[str]) -> bool:
    for prefix in azs_prefixes:
        if tr.text.find(prefix) != -1:
            return True
    return False


def parse_report(file_name: Path) -> dict:
    azs_prefixes = {'ALK', 'CHO', 'EKA', 'HMO', 'IVO', 'KDK', 'KEO', 'KYK', 'MSK', 'NNO', 'NSO',
                    'OMO', 'SPB', 'SVO', 'TUO', 'YAR', 'IAR', 'KRR', 'MOW', 'MSK', 'NN', 'RYZ',
                    'CEK', 'MSK', 'NN', 'SPB', 'TUO', 'YAR'}
    with open(str(file_name), encoding='utf8') as file:
        soup = BeautifulSoup(file.read(), "html.parser")

    tbody = soup.find("table")
    trs = tbody.find_all("tr")
    interesting_tr = []

    for tr in trs:
        if is_interesting(tr, azs_prefixes):
            interesting_tr.append(tr)

    data = {}
    for entry in interesting_tr:
        tds = entry.find_all("td")
        player_name = tds[1].text.strip('\n')
        for azs_prefix in azs_prefixes:
            try:
                azs_number = int(re.findall(rf'(?<={azs_prefix}[_-])\d+', player_name)[0])
                if not data.get(azs_number):
                    data[azs_number] = int(tds[5].text)
                else:
                    data[azs_number] += int(tds[5].text)
            except Exception as e:
                pass

    return data


def process_files(field1: Entry, field2: Entry) -> None:
    report_file_names = [Path(file_name) for file_name in field1.get().split(';')]
    template_file_name = Path(field2.get())

    report = {}
    for file_name in report_file_names:
        report.update(parse_report(file_name))

    wb = load_workbook(str(template_file_name))
    ws = wb.active

    flag = False
    for row in ws.iter_rows():
        if row[1].value == "№ АЗС:":
            flag = True
            continue
        elif row[1].value is None:
            flag = False
        if not flag:
            continue

        azs_code = row[1].value
        try:
            row[15].value = report[azs_code]
        except KeyError:
            row[15].value = 0

    wb.save(str(template_file_name))
    sys.exit()


if __name__ == "__main__":
    root = Tk()

    Label(root, text="Отчёт (.html): ").grid(row=0)
    Label(root, text="Шаблон (.xlsx): ").grid(row=1)

    text_field1 = Entry(root, width=60)
    text_field2 = Entry(root, width=60)

    text_field1.grid(row=0, column=1)
    text_field2.grid(row=1, column=1)

    b1_title = "Выбрать отчёт"
    b2_title = "Выбрать шаблон"
    b1 = Button(text=b1_title,
                command=lambda: select_report(title=b1_title, element=text_field1)
                )
    b2 = Button(text=b2_title,
                command=lambda: select_template(title=b2_title, element=text_field2)
                )
    b3 = Button(text="Run", command=lambda: process_files(text_field1, text_field2))

    b1.grid(row=0, column=2)
    b2.grid(row=1, column=2)
    b3.grid(row=2, column=1)

    mainloop()
