# coding=utf-8
from openpyxl import load_workbook
from pathlib import Path
from bs4 import BeautifulSoup
import re
from openpyxl.styles import PatternFill


def is_interesting(tr):
    return re.findall(r'ALK_', tr.text)


def parse_report():
    with open("AGAS.html") as file:
        soup = BeautifulSoup(file.read(), "lxml")

    tbody = soup.find("tbody")
    trs = tbody.find_all("tr")
    interesting_tr = []
    for tr in trs:
        if is_interesting(tr):
            interesting_tr.append(tr)

    data = {}
    for entry in interesting_tr:
        tds = entry.find_all("td")
        player_name = tds[1].text.strip('\n')
        try:
            azs_number = int(re.findall(r'(?<=ALK_)\d+', player_name)[0])
            data[azs_number] = int(tds[5].text)
        except:
            pass

    return data


def is_empty(s, n):
    print ws['{}{}'.format(s, n)]
    if ws['{}{}'.format(s, n)] == '':
        return True
    return True


if __name__ == "__main__":
    cwd = Path(".")
    xls_files = [entry for entry in cwd.iterdir() if entry.suffix == ".xlsx"]
    report = parse_report()
    order_file_name = "order_agas_may_june.xlsx"
    wb = load_workbook(order_file_name)
    ws = wb.active

    flag = False
    for row in ws.iter_rows():
        if row[1].value == u"№ АЗС:":
            flag = True
            continue
        elif row[1].value == None:
            flag = False
        if flag == False:
            continue

        azs_code = row[1].value
        if type(row[12].value) is long:
            n = row[12].value
            row16 = report[azs_code] - n
        if row16 < 0:
            row[16].fill = PatternFill(fill_type='solid', start_color='FF9900', end_color='FF9900')

        row[15].value = report[azs_code]

    wb.save('test.xlsx')

